import streamlit as st
import pandas as pd
import yfinance as yf
import re
import io
import os 
import altair as alt  # 新增：用來繪製可以精準控制排序的圖表
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.chart import BarChart, Reference

# ==========================================
# 網頁基本設定 (設定為寬螢幕)
# ==========================================
st.set_page_config(page_title="台灣50中100分價量試分析", layout="wide", page_icon="📈")

# ==========================================
# 狀態管理 (Session State) 初始化
# 用來記錄當前使用者點擊了哪一檔股票
# ==========================================
if 'selected_stock' not in st.session_state:
    st.session_state.selected_stock = None

# ==========================================
# 台股跳動點與區間函式
# ==========================================
def get_tick_size(price):
    if price < 10:
        return 0.01
    elif price < 50:
        return 0.05
    elif price < 100:
        return 0.1
    elif price < 500:
        return 0.5
    elif price < 1000:
        return 1.0
    else:
        return 5.0

def generate_ticks(low, high):
    low = round(low, 2)
    high = round(high, 2)
    if low >= high:
        return [low]
    ticks = []
    curr = low
    while curr <= high:
        ticks.append(curr)
        curr = round(curr + get_tick_size(curr), 2)
    if ticks[-1] < high:
        ticks.append(high)
    return list(dict.fromkeys(ticks))

# ==========================================
# 主程式邏輯封裝
# ==========================================
def run_analysis(df_excel):
    col_ticker = df_excel.columns[0]
    col_name = df_excel.columns[1]
    
    results = []
    serial_num = 1
    total_stocks = len(df_excel)
    progress_bar = st.progress(0)
    
    for row_idx, row in df_excel.iterrows():
        progress_bar.progress((row_idx + 1) / total_stocks)
        
        if pd.isna(row[col_ticker]) or str(row[col_ticker]).strip().lower() == 'nan':
            continue
            
        raw_ticker = str(row[col_ticker]).strip()
        stock_name = str(row[col_name]).strip() if pd.notna(row[col_name]) else "未知名稱"
        
        if raw_ticker.endswith('.0'):
            raw_ticker = raw_ticker[:-2]
            
        yf_ticker = f"{raw_ticker}.TW"
        
        try:
            stock_data = yf.Ticker(yf_ticker)
            hist = stock_data.history(period="6mo")
            
            if hist.empty:
                yf_ticker_two = f"{raw_ticker}.TWO"
                stock_data = yf.Ticker(yf_ticker_two)
                hist = stock_data.history(period="6mo")
                yf_ticker = yf_ticker_two
                
            if hist.empty:
                continue
                
            hist_64 = hist.tail(64).copy()
            current_price = hist_64['Close'].dropna().iloc[-1]
            current_price_round = round(current_price, 2)
            
            max_price = hist_64['High'].max()
            min_price = hist_64['Low'].min()
            
            if max_price == min_price:
                max_price = min_price * 1.05
                min_price = min_price * 0.95
                
            bin_size = (max_price - min_price) / 20
            
            if current_price_round >= max_price:
                curr_bin_idx = 19
            elif current_price_round <= min_price:
                curr_bin_idx = 0
            else:
                curr_bin_idx = int((current_price_round - min_price) / bin_size)
                if curr_bin_idx > 19: 
                    curr_bin_idx = 19
                    
            bins_data = []
            for i in range(20):
                start = min_price + i * bin_size
                end = min_price + (i + 1) * bin_size
                bins_data.append({'idx': i, 'start': start, 'end': end, 'vol': 0})
                
            price_vol_dict = {}
            for _, row_data in hist_64.iterrows():
                open_p = round(row_data['Open'], 2)
                high_p = round(row_data['High'], 2)
                low_p = round(row_data['Low'], 2)
                close_p = round(row_data['Close'], 2)
                vol = row_data['Volume']
                
                if vol == 0 or pd.isna(close_p):
                    continue
                    
                vol_close = vol * 0.30
                vol_open = vol * 0.05
                vol_remain = vol * 0.65
                
                price_vol_dict[close_p] = price_vol_dict.get(close_p, 0) + vol_close
                price_vol_dict[open_p] = price_vol_dict.get(open_p, 0) + vol_open
                
                ticks = generate_ticks(low_p, high_p)
                num_ticks = len(ticks)
                
                if num_ticks > 0:
                    vol_per_tick = vol_remain / num_ticks
                    for t in ticks:
                        price_vol_dict[t] = price_vol_dict.get(t, 0) + vol_per_tick
                        
            price_vol = pd.Series(price_vol_dict)
            
            for price, vol in price_vol.items():
                if price >= max_price:
                    bins_data[-1]['vol'] += vol
                elif price <= min_price:
                    bins_data[0]['vol'] += vol
                else:
                    idx = int((price - min_price) / bin_size)
                    if idx > 19: idx = 19
                    bins_data[idx]['vol'] += vol
                    
            valid_bins = [b for b in bins_data if b['vol'] > 0]
            top_3_bins = sorted(valid_bins, key=lambda x: x['vol'], reverse=True)[:3]
            
            for rank, t_bin in enumerate(top_3_bins, 1):
                if t_bin['idx'] - 3 <= curr_bin_idx <= t_bin['idx'] + 3:
                    safe_sheet_name = re.sub(r'[\\/*?:"<>|]', '_', stock_name)[:30]
                    disp_bins = []
                    for b in bins_data:  
                        is_current = (b['idx'] == curr_bin_idx)
                        prefix = "🎯 " if is_current else ""
                        disp_bins.append({
                            '區間標籤': f"{prefix}{b['start']:.2f} ~ {b['end']:.2f}",
                            '累積成交量': int(b['vol'])
                        })
                    
                    # 💡 【修改點 3 & 4】：將陣列反轉，確保價格高的在第一筆 (最上面)
                    disp_bins.reverse()

                    results.append({
                        '序號': serial_num,
                        '代碼': raw_ticker,
                        '個股名稱': stock_name,
                        'safe_name': safe_sheet_name,
                        '當日現價': current_price_round,
                        '落點分價': f"第{rank}大量",
                        '分價範圍': f"{t_bin['start']:.2f} ~ {t_bin['end']:.2f}",
                        'bins_data': disp_bins 
                    })
                    serial_num += 1
                    break
                    
        except Exception:
            continue
            
    progress_bar.empty()
    return results

# ==========================================
# 產生 Excel 檔案至記憶體 (供下載)
# ==========================================
def generate_excel(results):
    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "分價量落點分析"
    
    today_disp_str = datetime.now().strftime("%Y/%m/%d")
    ws_sum.append([f"分析日期: {today_disp_str}"])
    ws_sum['A1'].font = Font(bold=True, size=12)
    
    headers = ["序號", "個股名稱", "當日現價", "落點分價", "分價範圍"]
    ws_sum.append(headers)
    
    for col in range(1, 6):
        cell = ws_sum.cell(row=2, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        
    for idx, row_data in enumerate(results):
        row_num = idx + 3
        ws_sum.cell(row=row_num, column=1, value=row_data['序號'])
        
        link_cell = ws_sum.cell(row=row_num, column=2, value=row_data['個股名稱'])
        link_cell.hyperlink = f"#'{row_data['safe_name']}'!A1"
        link_cell.style = "Hyperlink" 
        
        ws_sum.cell(row=row_num, column=3, value=row_data['當日現價'])
        ws_sum.cell(row=row_num, column=4, value=row_data['落點分價'])
        ws_sum.cell(row=row_num, column=5, value=row_data['分價範圍'])
        
        ws_stock = wb.create_sheet(title=row_data['safe_name'])
        ws_stock.append(headers)
        ws_stock.append([row_data['序號'], row_data['個股名稱'], row_data['當日現價'], row_data['落點分價'], row_data['分價範圍']])
        
        for col in range(1, 6):
            ws_stock.cell(row=1, column=col).font = Font(bold=True)
        
        ws_stock.cell(row=5, column=1, value="價格級距區間 (TWD)").font = Font(bold=True)
        ws_stock.cell(row=5, column=2, value="累積成交量 (股)").font = Font(bold=True)
        
        for b_idx, b_data in enumerate(row_data['bins_data']):
            r = 6 + b_idx
            ws_stock.cell(row=r, column=1, value=b_data['區間標籤'])
            ws_stock.cell(row=r, column=2, value=b_data['累積成交量'])
            
        chart = BarChart()
        chart.type = "bar" 
        chart.style = 10
        chart.title = f"{row_data['個股名稱']} 64日分價量分佈圖"
        chart.x_axis.title = "價格級距區間 (TWD)"
        chart.y_axis.title = "累積成交量 (股)"
        chart.legend = None 
        chart.height = 14
        chart.width = 20
        
        data = Reference(ws_stock, min_col=2, min_row=5, max_row=25)
        cats = Reference(ws_stock, min_col=1, min_row=6, max_row=25)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws_stock.add_chart(chart, "D5")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ==========================================
# Streamlit UI 介面 
# ==========================================
st.title("📈 台灣50中100分價量試分析")
st.markdown("系統已設定為自動抓取近 64 日資料，將成交量依據台股跳動點精準均分，並找出收盤價位於「前三大成交量區間 ± 3 個抽屜」內的強勢/關鍵個股。")

file_path = 'TW50100.xlsx'

if os.path.exists(file_path):
    try:
        df_excel = pd.read_excel(file_path, engine='openpyxl', dtype=str)
        st.success(f"✅ 已成功自動載入清單：`{file_path}` (共 {len(df_excel)} 檔個股)")
        
        # 建立一個佔位符來儲存分析結果
        if 'analysis_results' not in st.session_state:
            st.session_state.analysis_results = None

        # 只有在尚未分析時才顯示開始分析按鈕
        if st.session_state.analysis_results is None:
            if st.button("🚀 開始分析", type="primary"):
                with st.spinner('連線抓取報價與分價量運算中，這可能需要幾分鐘的時間，請稍候...'):
                    st.session_state.analysis_results = run_analysis(df_excel)
                    st.rerun()  # 分析完畢後重新刷新網頁以隱藏按鈕
        
        results = st.session_state.analysis_results
        
        if results is not None:
            # 判斷目前要顯示哪一個畫面 (總覽表 or 獨立個股圖表)
            if st.session_state.selected_stock is None:
                
                # ==========================================
                # 視圖 1：總覽表畫面
                # ==========================================
                if results:
                    st.success(f"✅ 分析完成！共篩選出 **{len(results)}** 檔符合條件的個股。")
                    
                    excel_data = generate_excel(results)
                    today_str = datetime.now().strftime("%Y%m%d")
                    st.download_button(
                        label="📥 下載完整含圖表的 Excel 報表",
                        data=excel_data,
                        file_name=f"分價每日分析_{today_str}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
                    st.subheader("📋 符合條件個股總覽")
                    st.markdown("💡 *點擊名稱按鈕，即可展開該檔股票的詳細分價圖表*")
                    
                    # 自製版面來呈現帶有按鈕的表格
                    col1, col2, col3, col4, col5 = st.columns([1, 2, 1.5, 1.5, 2])
                    col1.markdown("**代碼**")
                    col2.markdown("**個股名稱**")
                    col3.markdown("**當日現價**")
                    col4.markdown("**命中條件**")
                    col5.markdown("**大量區範圍**")
                    st.divider()
                    
                    for r in results:
                        c1, c2, c3, c4, c5 = st.columns([1, 2, 1.5, 1.5, 2])
                        c1.write(r['代碼'])
                        
                        # 按下按鈕時，將選擇的股票寫入 session state 並重整畫面
                        if c2.button(f"🔍 {r['個股名稱']}", key=f"btn_{r['代碼']}", help="點擊查看詳細圖表", type="secondary"):
                            st.session_state.selected_stock = r['代碼']
                            st.rerun()
                            
                        c3.write(r['當日現價'])
                        c4.write(r['落點分價'])
                        c5.write(r['分價範圍'])
                else:
                    st.warning("📊 分析完成，今日無符合條件的個股。")
            
            else:
                # ==========================================
                # 視圖 2：單一個股詳細圖表畫面
                # ==========================================
                # 找出被點擊的那檔股票資料
                target_r = next((r for r in results if r['代碼'] == st.session_state.selected_stock), None)
                
                if target_r:
                    # 返回按鈕
                    if st.button("🔙 返回符合條件個股總覽", type="primary"):
                        st.session_state.selected_stock = None
                        st.rerun()
                        
                    st.subheader(f"📌 {target_r['代碼']} {target_r['個股名稱']} (現價: {target_r['當日現價']})")
                    st.markdown(f"**符合條件：** 落在 {target_r['落點分價']}")
                    
                    col1, col2 = st.columns([1, 2])
                    
                    # 轉換 bins 資料
                    df_bins = pd.DataFrame(target_r['bins_data'])
                    
                    with col1:
                        st.markdown("**價格級距資料表 (🎯 標示為現價區間)**")
                        # 將區間標籤設為索引來顯示乾淨的資料表，高價已在第一列
                        st.dataframe(df_bins.set_index('區間標籤'), use_container_width=True)
                        
                    with col2:
                        st.markdown("**64日分價量分佈長條圖**")
                        # 擷取順序陣列，用來強迫 Altair 依照「高價在上、低價在下」渲染 Y 軸
                        sort_order = df_bins['區間標籤'].tolist()
                        
                        chart = alt.Chart(df_bins).mark_bar(color='#1f77b4').encode(
                            x=alt.X('累積成交量:Q', title='累積成交量 (股)'),
                            # Y 軸套用自訂的降冪順序 (sort=sort_order)
                            y=alt.Y('區間標籤:N', title='價格級距區間 (TWD)', sort=sort_order),
                            tooltip=['區間標籤', '累積成交量']
                        ).properties(height=550)  # 稍微拉高圖表以符合 20 個抽屜的視覺
                        
                        st.altair_chart(chart, use_container_width=True)

    except Exception as e:
        st.error(f"❌ 發生錯誤：{e}")
        
else:
    st.error(f"❌ 找不到股票清單檔案：`{file_path}`")
    st.info("💡 如果在您的電腦上執行，請確認 `TW50100.xlsx` 有跟 `app.py` 放在同一個資料夾。")
